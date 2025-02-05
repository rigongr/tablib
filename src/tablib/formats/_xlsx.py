""" Tablib - XLSX Support.
"""

import re
from io import BytesIO

from openpyxl.reader.excel import ExcelReader, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

import tablib as tablib

INVALID_TITLE_REGEX = re.compile(r'[\\*?:/\[\]]')

def safe_xlsx_sheet_title(s, replace="-"):
    return re.sub(INVALID_TITLE_REGEX, replace, s)[:31]


class XLSXFormat:
    title = 'xlsx'
    extensions = ('xlsx',)

    @classmethod
    def detect(cls, stream):
        """Returns True if given stream is a readable excel file."""
        try:
            # No need to fully load the file, it should be enough to be able to
            # read the manifest.
            reader = ExcelReader(stream, read_only=False)
            reader.read_manifest()
            return True
        except Exception:
            return False

    @classmethod
    def export_set(cls, dataset, freeze_panes=True, invalid_char_subst="-", something=None):
        """Returns XLSX representation of Dataset.

        If dataset.title contains characters which are considered invalid for an XLSX file
        sheet name (http://www.excelcodex.com/2012/06/worksheets-naming-conventions/), they will
        be replaced with `invalid_char_subst`.
        """
        wb = Workbook()
        ws = wb.worksheets[0]
        breakpoint()
        ws.title = (
            safe_xlsx_sheet_title(dataset.title, invalid_char_subst)
            if dataset.title else 'Tablib Dataset'
        )

        cls.dset_sheet(dataset, ws, freeze_panes=freeze_panes)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @classmethod
    def export_book(cls, databook, freeze_panes=True, invalid_char_subst="-"):
        """Returns XLSX representation of DataBook.
        
        If dataset.title contains characters which are considered invalid for an XLSX file
        sheet name (http://www.excelcodex.com/2012/06/worksheets-naming-conventions/), they will
        be replaced with `invalid_char_subst`.
        """
        print('in here')
        breakpoint()
        wb = Workbook()
        for sheet in wb.worksheets:
            wb.remove(sheet)
        for i, dset in enumerate(databook._datasets):
            ws = wb.create_sheet()
            ws.title = (
                safe_xlsx_sheet_title(dset.title, invalid_char_subst)
                if dset.title else 'Sheet%s' % (i)
            )

            cls.dset_sheet(dset, ws, freeze_panes=freeze_panes)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @classmethod
    def import_set(cls, dset, in_stream, headers=True, read_only=True, skip_lines=0):
        """Returns databook from XLS stream."""

        dset.wipe()

        xls_book = load_workbook(in_stream, read_only=read_only, data_only=True)
        sheet = xls_book.active

        dset.title = sheet.title

        for i, row in enumerate(sheet.rows):
            if i < skip_lines:
                continue
            row_vals = [c.value for c in row]
            if i == skip_lines and headers:
                dset.headers = row_vals
            else:
                dset.append(row_vals)

    @classmethod
    def import_book(cls, dbook, in_stream, headers=True, read_only=True):
        """Returns databook from XLS stream."""

        dbook.wipe()

        xls_book = load_workbook(in_stream, read_only=read_only, data_only=True)

        for sheet in xls_book.worksheets:
            data = tablib.Dataset()
            data.title = sheet.title

            for i, row in enumerate(sheet.rows):
                row_vals = [c.value for c in row]
                if (i == 0) and (headers):
                    data.headers = row_vals
                else:
                    if i > 0 and len(row_vals) < data.width:
                        row_vals += [''] * (data.width - len(row_vals))
                    data.append(row_vals)

            dbook.add_sheet(data)

    @classmethod
    def dset_sheet(cls, dataset, ws, freeze_panes=True):
        """Completes given worksheet from given Dataset."""
        _package = dataset._package(dicts=False)

        for i, sep in enumerate(dataset._separators):
            _offset = i
            _package.insert((sep[0] + _offset), (sep[1],))

        bold = Font(bold=True, 
                    color='00FFFFFF',
                    
                )

        wrap_text = Alignment(wrap_text=True)

        for i, row in enumerate(_package):
            row_number = i + 1
            for j, col in enumerate(row):
                col_idx = get_column_letter(j + 1)
                cell = ws[f'{col_idx}{row_number}']
                cell.font = Font(size=10, name='Arial')
                cell.alignment = Alignment(horizontal='left', vertical='top')
                # bold headers
                if (row_number == 1) and dataset.headers:
                    cell.font = bold
                    cell.fill = PatternFill("solid", bgColor='00000000')
                    if freeze_panes:
                        #  Export Freeze only after first Line
                        ws.freeze_panes = 'A2'

                # bold separators
                elif len(row) < dataset.width:
                    cell.font = bold

                # wrap the rest
                else:
                    try:
                        str_col_value = str(col)
                    except TypeError:
                        str_col_value = ''
                    if '\n' in str_col_value:
                        cell.alignment = wrap_text

                try:
                    cell.value = col
                except (ValueError, TypeError):
                    cell.value = str(col)
